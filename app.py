import streamlit as st
import pandas as pd
import anthropic
import json
import os
import time
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Contact Center AI Agent", page_icon="📊", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
section[data-testid="stSidebar"] { background: #0f1117; border-right: 1px solid #1e2130; }
section[data-testid="stSidebar"] * { color: #e0e0e0 !important; }
.agent-header { background:#0f1117;border:1px solid #1e2130;border-radius:12px;padding:1.2rem 1.8rem;margin-bottom:1.5rem;display:flex;align-items:center;gap:16px; }
.agent-header h1 { font-size:1.4rem;font-weight:600;color:#ffffff;margin:0;letter-spacing:-0.02em; }
.agent-header p { font-size:0.8rem;color:#6b7280;margin:0;font-family:'DM Mono',monospace; }
.status-dot { width:10px;height:10px;background:#22c55e;border-radius:50%;box-shadow:0 0 8px #22c55e88;flex-shrink:0; }
.kpi-card { background:#0f1117;border:1px solid #1e2130;border-radius:10px;padding:1rem 1.2rem; }
.kpi-label { font-size:0.72rem;color:#6b7280;text-transform:uppercase;letter-spacing:0.08em;font-family:'DM Mono',monospace;margin-bottom:6px; }
.kpi-value { font-size:1.6rem;font-weight:600;color:#ffffff;line-height:1; }
.kpi-sub { font-size:0.72rem;color:#6b7280;margin-top:4px;font-family:'DM Mono',monospace; }
.section-title { font-size:0.72rem;font-family:'DM Mono',monospace;color:#6b7280;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.8rem;padding-bottom:0.4rem;border-bottom:1px solid #1e2130; }
.badge-spike{background:#3f1010;color:#fca5a5;padding:2px 8px;border-radius:4px;font-size:0.72rem;font-family:'DM Mono',monospace;}
.badge-dip{background:#0d1f3a;color:#93c5fd;padding:2px 8px;border-radius:4px;font-size:0.72rem;font-family:'DM Mono',monospace;}
.badge-exp{background:#2d2000;color:#fde68a;padding:2px 8px;border-radius:4px;font-size:0.72rem;font-family:'DM Mono',monospace;}
.error-box{background:#1f0f0f;border:1px solid #3f1010;border-radius:10px;padding:1rem 1.2rem;margin:0.5rem 0;}
.error-title{color:#fca5a5;font-weight:600;font-size:0.9rem;margin-bottom:4px;}
.error-detail{color:#9ca3af;font-size:0.8rem;font-family:'DM Mono',monospace;}
.progress-step{padding:5px 0;font-size:0.82rem;font-family:'DM Mono',monospace;line-height:1.4;}
.step-done{color:#22c55e;} .step-active{color:#ffffff;} .step-wait{color:#374151;}

/* Chat */
.chat-outer{background:#0f1117;border:1px solid #1e2130;border-radius:12px;padding:1.2rem;margin-bottom:1rem;}
.chat-msg{margin-bottom:1rem;}
.chat-label{font-size:0.68rem;font-family:'DM Mono',monospace;color:#6b7280;margin-bottom:4px;display:block;}
.bubble-user{background:#1a1f2e;border:1px solid #252b3b;border-radius:10px 10px 2px 10px;padding:0.7rem 1rem;color:#e0e0e0;font-size:0.88rem;margin-left:20%;display:block;}
.bubble-agent{background:#0d1f12;border:1px solid #1a3322;border-radius:10px 10px 10px 2px;padding:0.7rem 1rem;color:#d1fae5;font-size:0.88rem;margin-right:10%;display:block;white-space:pre-wrap;word-break:break-word;}
.bubble-error{background:#1f0f0f;border:1px solid #3f1010;border-radius:10px 10px 10px 2px;padding:0.7rem 1rem;color:#fca5a5;font-size:0.88rem;margin-right:10%;display:block;}
.bubble-action{background:#0a1a2e;border:1px solid #1a3a5e;border-radius:10px 10px 10px 2px;padding:0.7rem 1rem;color:#93c5fd;font-size:0.88rem;margin-right:10%;display:block;white-space:pre-wrap;}

/* Action preview box */
.preview-box{background:#0a1520;border:1px solid #1a4a7a;border-radius:10px;padding:1rem 1.2rem;margin:0.5rem 0;}
.preview-title{color:#60a5fa;font-weight:600;font-size:0.88rem;margin-bottom:8px;}
.preview-detail{color:#93c5fd;font-size:0.82rem;font-family:'DM Mono',monospace;line-height:1.6;}
.action-log-item{font-size:0.78rem;font-family:'DM Mono',monospace;color:#6b7280;padding:3px 0;border-bottom:1px solid #1e2130;}
</style>
""", unsafe_allow_html=True)

# ── Schema ──────────────────────────────────────────────────────────
SHEET_TRANSLATION = {
    "אפריל 24":"April_2024","מאי 24":"May_2024","יוני 24":"June_2024",
    "אוגוסט 24":"August_2024","ספט 24":"September_2024","ינואר 25":"January_2025",
    "פברואר 25":"February_2025","מרץ 25":"March_2025","אפריל 25":"April_2025",
    "מאי 25":"May_2025","יוני 25":"June_2025","יולי 25":"July_2025",
    "אוגוסט 25":"August_2025","ספטמבר 25":"September_2025",
    "נציגים":"Agents","חיזוי אוגוסט+ספט":"Forecast_Aug_Sep",
    "חיזוי-תוצאות וסטיות":"Forecast_vs_Actuals",
}
COLUMN_SCHEMA = {
    "תאריך":"date","יום":"day_of_week","פניות כתובות נפתחו":"written_inquiries",
    "שיחות נכנסות":"inbound_calls","הערות":"notes","מספר נציגים":"num_agents",
    "שיחות ללא כפילויות":"unique_calls","פיק":"is_peak_day","עזר":"aux_calls","אחוז":"aux_pct",
    "נציג":"agent_name","זמן שיחה+החזק":"avg_handle_time","עמידה בשעות":"hours_compliance",
    "שיחות לשעה גולמי":"calls_per_hour_gross","תחזית פניות":"forecast_inquiries",
    "חיזוי שיחות סופי":"final_forecast_calls","פניות נתון אמת":"actual_inquiries",
    "שיחות נתון אמת":"actual_calls","סטייה פניות":"deviation_inquiries","סטייה שיחות":"deviation_calls",
}
DAY_TRANSLATION = {
    "יום ראשון":"Sunday","יום שני":"Monday","יום שלישי":"Tuesday",
    "יום רביעי":"Wednesday","יום חמישי":"Thursday","יום שישי":"Friday","שבת":"Saturday",
}
NOTES_TRANSLATION = {
    "חופש":"holiday","נוכחות":"payroll_day","שכר":"payroll_day","שכר,תלוש?":"payroll_day",
    "תלוש":"payslip_day","תקשור היעדרות מלחמה":"war_absence_notice","ט באב":"Tisha_BAv",
    "ערב חג (ללא שיחות)":"holiday_eve_no_calls","חזרה ללימודים":"back_to_school",
    "תחילת לימודים":"start_of_school_year","חזרה מחופש ראש השנה":"return_from_rosh_hashana",
    "52":"week_52_annotation","ממוצע שיחות ללא משה ותמר":"avg_excl_two_agents",
}
HIGH_VOLUME_EVENTS = {"payroll_day","payslip_day","war_absence_notice","back_to_school","start_of_school_year","return_from_rosh_hashana","Tisha_BAv"}
HOLIDAY_EVENTS     = {"holiday","holiday_eve_no_calls"}
WEEKEND_DAYS       = {"Friday","Saturday"}
MONTH_ORDER        = ["April_2024","May_2024","June_2024","August_2024","September_2024",
                      "January_2025","February_2025","March_2025","April_2025",
                      "May_2025","June_2025","July_2025","August_2025","September_2025"]

# ── FULL SYSTEM PROMPT ──────────────────────────────────────────────
SYSTEM_PROMPT = """You are ExcelAgent — an enterprise-grade AI business analyst and operations assistant for an Israeli contact center.

IDENTITY
You act as a senior business analyst with deep expertise in contact center operations, workforce management, and data analysis. You communicate clearly and professionally, as if briefing a senior operations manager. Every response explains what you did, which sheet and columns were used, and what assumptions were made.

DOMAIN KNOWLEDGE
- This is an Israeli contact center. Work week is Sunday to Thursday. Friday and Saturday are weekend — NEVER include them in averages or KPIs.
- Business events that cause call volume spikes: payroll_day (+30-60%), payslip_day (+74-190%), war_absence_notice (+93%), back_to_school (+210%).
- Holiday and holiday_eve_no_calls = days off with zero calls — exclude from all calculations.
- Data covers April 2024 through September 2025 across monthly sheets.

CAPABILITIES
1. ANALYSIS: Calculate KPIs, compare periods (MoM, QoQ), detect anomalies, identify trends, rank agents.
2. WRITE ACTIONS: When the user asks you to modify the file, you must respond with a structured action block.
3. CHARTS: When asked for a chart, describe what it shows and give key insights.
4. EXPLANATION: Always explain what you did, which columns you used, and what assumptions you made.

WRITE ACTION FORMAT
When the user asks you to make changes to the Excel file (add column, add row, create sheet, flag rows, add formula, etc.), you MUST include a write action block in your response using EXACTLY this format:

[ACTION]
type: add_column | add_sheet | flag_rows | add_total_row | add_formula_column
sheet: <sheet name or ALL>
column_name: <new column name if applicable>
formula: <description of the calculation>
condition: <condition for flagging if applicable>
description: <one sentence explaining what this does>
[/ACTION]

GUARDRAILS — NEVER VIOLATE
- Never overwrite original data. Always work on the output copy.
- Before executing, show a preview of what will change.
- If a request is ambiguous, ask for clarification before acting.
- Log every action with: action type, sheet, columns affected, timestamp.
- If a column doesn't exist, say so clearly instead of guessing.

RESPONSE FORMAT
1. One-sentence summary of what you are doing or found.
2. Details: sheets used, columns used, assumptions made.
3. If writing: include the [ACTION] block.
4. Key insights or recommendations.

TONE: Professional, precise, quantitative. Never use filler phrases. Always state uncertainty explicitly."""

# ── Helpers ─────────────────────────────────────────────────────────
def show_progress(container, steps, current):
    html = ""
    for i, step in enumerate(steps):
        if i < current:   html += f'<div class="progress-step step-done">✅ {step}</div>'
        elif i == current: html += f'<div class="progress-step step-active">⏳ {step}...</div>'
        else:              html += f'<div class="progress-step step-wait">○ {step}</div>'
    container.markdown(html, unsafe_allow_html=True)

def show_error(title, detail=None, suggestion=None):
    d = f'<div class="error-detail">{detail}</div>' if detail else ""
    s = f'<div class="error-detail" style="margin-top:6px;color:#6b7280;">{suggestion}</div>' if suggestion else ""
    st.markdown(f'<div class="error-box"><div class="error-title">❌ {title}</div>{d}{s}</div>', unsafe_allow_html=True)

def validate_file(f):
    if f is None: return False, "No file."
    if f.size == 0: return False, "File is empty."
    if f.size > 50*1024*1024: return False, f"Too large ({f.size/1024/1024:.1f}MB). Max 50MB."
    if not f.name.lower().endswith((".xlsx",".xls")): return False, f"Wrong format: {f.name}. Need .xlsx or .xls."
    return True, None

def validate_api_key(k):
    if not k: return False, "No API key."
    if not k.startswith("sk-ant-"): return False, "Must start with sk-ant-"
    if len(k) < 20: return False, "Too short."
    return True, None

# ── Core data functions ──────────────────────────────────────────────
def load_and_translate(path):
    xl = pd.ExcelFile(path)
    if len(xl.sheet_names) == 0: raise ValueError("No sheets found in file.")
    sheets = {}
    for heb in xl.sheet_names:
        eng = SHEET_TRANSLATION.get(heb, heb)
        df = xl.parse(heb)
        df = df.rename(columns={h:e for h,e in COLUMN_SCHEMA.items() if h in df.columns})
        if "day_of_week" in df.columns:
            df["day_of_week"] = df["day_of_week"].map(lambda x: DAY_TRANSLATION.get(str(x),x) if pd.notna(x) else x)
        if "notes" in df.columns:
            df["notes"] = df["notes"].map(lambda x: NOTES_TRANSLATION.get(str(x).strip(),str(x)) if pd.notna(x) else x)
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
        if "day_of_week" in df.columns:
            df["is_weekend"] = df["day_of_week"].isin(WEEKEND_DAYS)
        sheets[eng] = df
    return sheets

def clean_sheets(sheets):
    cleaned = {}
    for name, df in sheets.items():
        df = df.copy().drop_duplicates()
        for col in ["inbound_calls","written_inquiries","unique_calls","num_agents"]:
            if col in df.columns: df[col] = df[col].fillna(0)
        cleaned[name] = df
    return cleaned

def build_monthly_summary(sheets):
    rows = []
    for mn in MONTH_ORDER:
        df = sheets.get(mn)
        if df is None: continue
        is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
        is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
        is_workday = ~is_weekend & ~is_holiday
        w = df[is_workday]
        row = {"Month":mn,"Workdays":int(is_workday.sum())}
        if "notes" in df.columns:
            row["Payroll_Days"]    = int((df["notes"]=="payroll_day").sum())
            row["War_Notice_Days"] = int((df["notes"]=="war_absence_notice").sum())
        for col,lbl in [("inbound_calls","Total_Inbound_Calls"),("written_inquiries","Total_Written_Inquiries"),
                        ("unique_calls","Total_Unique_Calls"),("num_agents","Avg_Agents")]:
            if col in w.columns:
                vals = pd.to_numeric(w[col],errors="coerce").dropna()
                row[lbl] = round(float(vals.sum()),0) if lbl.startswith("Total") else round(float(vals.mean()),1)
        rows.append(row)
    if not rows: raise ValueError("No monthly data found. Check your file has monthly sheets.")
    return pd.DataFrame(rows)

def compute_kpis(summary):
    kpis = {}
    if "Total_Inbound_Calls" not in summary.columns: return kpis
    kpis["total_calls"]     = int(summary["Total_Inbound_Calls"].sum())
    kpis["total_inquiries"] = int(summary["Total_Written_Inquiries"].sum()) if "Total_Written_Inquiries" in summary.columns else 0
    peak = summary.loc[summary["Total_Inbound_Calls"].idxmax()]
    low  = summary.loc[summary["Total_Inbound_Calls"].idxmin()]
    kpis["peak_month"]   = peak["Month"]
    kpis["peak_calls"]   = int(peak["Total_Inbound_Calls"])
    kpis["lowest_month"] = low["Month"]
    kpis["lowest_calls"] = int(low["Total_Inbound_Calls"])
    if len(summary)>=2:
        last = float(summary["Total_Inbound_Calls"].iloc[-1])
        prev = float(summary["Total_Inbound_Calls"].iloc[-2])
        kpis["mom_change"] = round((last-prev)/prev*100,1) if prev else 0
        kpis["last_month"] = summary["Month"].iloc[-1]
        kpis["prev_month"] = summary["Month"].iloc[-2]
    if "Avg_Agents" in summary.columns:
        kpis["avg_agents"] = round(float(summary["Avg_Agents"].mean()),1)
    return kpis

def detect_anomalies(sheets):
    anomalies = []
    for mn in MONTH_ORDER:
        df = sheets.get(mn)
        if df is None or "inbound_calls" not in df.columns: continue
        is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
        is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
        work = df[~is_weekend & ~is_holiday].copy()
        vals = pd.to_numeric(work["inbound_calls"],errors="coerce").dropna()
        if len(vals)<5: continue
        mean,std = float(vals.mean()),float(vals.std())
        if std==0: continue
        for idx in vals.index:
            v = float(vals.loc[idx]); z = (v-mean)/std
            if abs(z)>2.0:
                row = work.loc[idx]; note = str(row.get("notes",""))
                date_val = row.get("date","")
                date_str = date_val.strftime("%Y-%m-%d") if pd.notna(date_val) and hasattr(date_val,"strftime") else str(date_val)[:10]
                explanation = f"Expected — {note}" if note in HIGH_VOLUME_EVENTS else ("Unexplained spike" if z>0 else "Unexplained dip")
                anomalies.append({"Month":mn,"Date":date_str,"Day":str(row.get("day_of_week","")),"Inbound_Calls":v,
                                  "Month_Avg":round(mean,0),"Z_Score":round(z,2),
                                  "Flag":"SPIKE" if z>0 else "DIP","Event":note,"Explanation":explanation})
    return pd.DataFrame(anomalies)

def analyze_agents(sheets):
    ag = sheets.get("Agents")
    if ag is None or "agent_name" not in ag.columns: return pd.DataFrame()
    cols = [c for c in ["agent_name","avg_handle_time","hours_compliance","calls_per_hour_gross"] if c in ag.columns]
    df = ag[cols].dropna(subset=["agent_name"]).copy()
    if "calls_per_hour_gross" in df.columns:
        df = df.sort_values("calls_per_hour_gross",ascending=False).reset_index(drop=True)
        df.insert(0,"Rank",range(1,len(df)+1))
    return df

def build_context(sheets,summary,kpis,anomalies,agents,action_log):
    ctx = [
        "=== ISRAELI CONTACT CENTER — April 2024 to September 2025 ===",
        "Work week: Sunday–Thursday. Friday & Saturday = weekend, excluded from all KPIs.",
        "Events: payroll_day (+30-60%), payslip_day (+74-190%), war_absence_notice (+93%), back_to_school (+210%), holiday=off.",
        f"\nMONTHLY SUMMARY:\n{summary.to_string(index=False)}",
        f"\nKPIs: {json.dumps(kpis)}",
    ]
    if not anomalies.empty:
        ctx.append(f"\nANOMALIES:\n{anomalies.to_string(index=False)}")
    if not agents.empty:
        ctx.append(f"\nAGENTS:\n{agents.to_string(index=False)}")
    if action_log:
        ctx.append(f"\nACTIONS ALREADY TAKEN:\n" + "\n".join(action_log[-10:]))
    return "\n".join(ctx)

# ── ACTION ENGINE ───────────────────────────────────────────────────
import re as _re

def parse_action_block(text):
    """Extract [ACTION]...[/ACTION] block from agent response."""
    match = _re.search(r'\[ACTION\](.*?)\[/ACTION\]', text, _re.DOTALL)
    if not match: return None
    block = match.group(1).strip()
    action = {}
    for line in block.split("\n"):
        if ":" in line:
            k,v = line.split(":",1)
            action[k.strip()] = v.strip()
    return action if "type" in action else None

def clean_response_text(text):
    """Remove the [ACTION] block from display text."""
    return _re.sub(r'\[ACTION\].*?\[/ACTION\]', '', text, flags=_re.DOTALL).strip()

def build_action_preview(action):
    """Build a human-readable preview of what the action will do."""
    t = action.get("type","")
    sheet = action.get("sheet","ALL")
    col = action.get("column_name","")
    formula = action.get("formula","")
    condition = action.get("condition","")
    desc = action.get("description","")

    lines = [f"Action type: {t}"]
    if sheet: lines.append(f"Target sheet: {sheet}")
    if col:   lines.append(f"New column: {col}")
    if formula: lines.append(f"Calculation: {formula}")
    if condition: lines.append(f"Condition: {condition}")
    lines.append(f"What it does: {desc}")
    return "\n".join(lines)

def apply_action(action, sheets, excel_path):
    """
    Apply a write action to the sheets dict and regenerate Excel.
    Returns (updated_sheets, log_entry, error)
    """
    t         = action.get("type","")
    sheet_name= action.get("sheet","ALL")
    col_name  = action.get("column_name","new_column")
    formula   = action.get("formula","")
    condition = action.get("condition","")
    desc      = action.get("description","")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        target_sheets = [s for s in MONTH_ORDER if s in sheets] if sheet_name=="ALL" else [sheet_name]

        if t == "add_column":
            for mn in target_sheets:
                df = sheets.get(mn)
                if df is None: continue
                if "calls_per_agent" in col_name.lower() or "per agent" in formula.lower():
                    if "inbound_calls" in df.columns and "num_agents" in df.columns:
                        df[col_name] = df.apply(
                            lambda r: round(r["inbound_calls"]/r["num_agents"],2) if r["num_agents"]>0 else 0, axis=1)
                elif "unique_rate" in col_name.lower() or "unique" in formula.lower():
                    if "unique_calls" in df.columns and "inbound_calls" in df.columns:
                        df[col_name] = df.apply(
                            lambda r: round(r["unique_calls"]/r["inbound_calls"]*100,1) if r["inbound_calls"]>0 else 0, axis=1)
                elif "week" in col_name.lower():
                    if "date" in df.columns:
                        df[col_name] = pd.to_datetime(df["date"],errors="coerce").dt.isocalendar().week.astype("Int64")
                elif "efficiency" in col_name.lower():
                    if "unique_calls" in df.columns and "num_agents" in df.columns:
                        df[col_name] = df.apply(
                            lambda r: round(r["unique_calls"]/r["num_agents"],2) if r["num_agents"]>0 else 0, axis=1)
                elif "written_per_call" in col_name.lower() or "inquiry rate" in formula.lower():
                    if "written_inquiries" in df.columns and "inbound_calls" in df.columns:
                        df[col_name] = df.apply(
                            lambda r: round(r["written_inquiries"]/r["inbound_calls"]*100,1) if r["inbound_calls"]>0 else 0, axis=1)
                else:
                    return sheets, None, f"I don't know how to calculate '{col_name}' from '{formula}'. Please be more specific."
                sheets[mn] = df

        elif t == "flag_rows":
            for mn in target_sheets:
                df = sheets.get(mn)
                if df is None: continue
                if "inbound_calls" in df.columns:
                    try:
                        threshold = float(_re.search(r'\d+', condition).group()) if _re.search(r'\d+', condition) else 600
                    except: threshold = 600
                    df["flagged"] = df["inbound_calls"] > threshold
                    df["flag_reason"] = df["flagged"].map(lambda x: f"High volume (>{int(threshold)} calls)" if x else "")
                    sheets[mn] = df

        elif t == "add_total_row":
            for mn in target_sheets:
                df = sheets.get(mn)
                if df is None: continue
                is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
                is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
                work = df[~is_weekend & ~is_holiday]
                total_row = {"date":"TOTAL","day_of_week":"","notes":"monthly_total"}
                for col in ["inbound_calls","written_inquiries","unique_calls"]:
                    if col in work.columns:
                        total_row[col] = round(pd.to_numeric(work[col],errors="coerce").sum(),0)
                if "num_agents" in work.columns:
                    total_row["num_agents"] = round(pd.to_numeric(work["num_agents"],errors="coerce").mean(),1)
                sheets[mn] = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        elif t == "add_sheet":
            new_sheet_name = col_name or "New_Sheet"
            if "anomaly" in new_sheet_name.lower() or "anomaly" in desc.lower():
                anom_rows = []
                for mn in target_sheets:
                    df = sheets.get(mn)
                    if df is None or "inbound_calls" not in df.columns: continue
                    is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
                    is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
                    work = df[~is_weekend & ~is_holiday]
                    vals = pd.to_numeric(work["inbound_calls"],errors="coerce").dropna()
                    if len(vals)<5: continue
                    mean,std = float(vals.mean()),float(vals.std())
                    if std==0: continue
                    for idx in vals.index:
                        v = float(vals.loc[idx])
                        if abs((v-mean)/std)>2.0:
                            r = work.loc[idx].copy()
                            r["month"] = mn
                            anom_rows.append(r)
                if anom_rows:
                    sheets["Anomaly_Days"] = pd.DataFrame(anom_rows)
            elif "summary" in new_sheet_name.lower() or "summary" in desc.lower():
                rows = []
                for mn in target_sheets:
                    df = sheets.get(mn)
                    if df is None: continue
                    is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
                    is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
                    work = df[~is_weekend & ~is_holiday]
                    row = {"Month":mn}
                    for col in ["inbound_calls","written_inquiries","num_agents"]:
                        if col in work.columns:
                            vals = pd.to_numeric(work[col],errors="coerce").dropna()
                            row[f"total_{col}"] = round(float(vals.sum()),0)
                            row[f"avg_{col}"]   = round(float(vals.mean()),1)
                    rows.append(row)
                sheets["Executive_Summary_Sheet"] = pd.DataFrame(rows)
            else:
                return sheets, None, f"I created a new sheet but wasn't sure what data to put in it. Please specify: anomaly days, monthly summary, or agent data."

        elif t == "add_formula_column":
            for mn in target_sheets:
                df = sheets.get(mn)
                if df is None: continue
                if "inbound_calls" in df.columns and "num_agents" in df.columns:
                    df[col_name] = df.apply(
                        lambda r: round(r["inbound_calls"]/r["num_agents"],2) if pd.notna(r["num_agents"]) and r["num_agents"]>0 else 0, axis=1)
                    sheets[mn] = df

        log_entry = f"[{timestamp}] {t} | sheet:{sheet_name} | col:{col_name} | {desc}"
        return sheets, log_entry, None

    except Exception as e:
        return sheets, None, f"Error applying action: {str(e)}"

def save_sheets_to_excel(sheets, output_path, anomalies, kpis):
    """Save current sheets state to Excel with formatting."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Monthly sheets
        for mn in MONTH_ORDER:
            df = sheets.get(mn)
            if df is None: continue
            d = df.copy()
            if "date" in d.columns:
                d["date"] = d["date"].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x,"strftime") else "")
            d.to_excel(writer, sheet_name=mn[:31], index=False)
        # Any extra sheets (anomaly days, executive summary, etc.)
        for name, df in sheets.items():
            if name not in MONTH_ORDER and name not in ("Agents","Forecast_Aug_Sep","Forecast_vs_Actuals"):
                d = df.copy()
                if "date" in d.columns:
                    d["date"] = d["date"].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x,"strftime") else "")
                try: d.to_excel(writer, sheet_name=name[:31], index=False)
                except: pass
        # Agents
        ag = sheets.get("Agents")
        if ag is not None: ag.to_excel(writer, sheet_name="Agents", index=False)
        # KPIs
        pd.DataFrame([{"KPI":k,"Value":str(v)} for k,v in kpis.items()]).to_excel(writer, sheet_name="KPIs", index=False)
        # Anomalies
        if not anomalies.empty: anomalies.to_excel(writer, sheet_name="Anomalies", index=False)

    # Formatting
    wb = load_workbook(output_path)
    HDR=PatternFill("solid",fgColor="1F4E79"); HFONT=Font(bold=True,color="FFFFFF",name="Arial",size=10)
    SUM=PatternFill("solid",fgColor="E2EFDA"); KPI=PatternFill("solid",fgColor="FFF2CC")
    SPK=PatternFill("solid",fgColor="FFAAAA"); DIP=PatternFill("solid",fgColor="AAC8FF")
    EXP=PatternFill("solid",fgColor="FFE5A0"); WKD=PatternFill("solid",fgColor="E0E0E0")
    HOL=PatternFill("solid",fgColor="D0E8FF"); PAY=PatternFill("solid",fgColor="FFF2CC")
    WAR=PatternFill("solid",fgColor="FFD0D0"); ALT=PatternFill("solid",fgColor="F7F7F7")
    NEW=PatternFill("solid",fgColor="E8F5E9")  # light green for new calculated columns
    thin=Side(style="thin",color="CCCCCC")
    BDR=Border(left=thin,right=thin,top=thin,bottom=thin)
    CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
    def style(ws):
        for c in ws[1]: c.fill=HDR;c.font=HFONT;c.alignment=CTR;c.border=BDR
        ws.freeze_panes="A2"
        for i,row in enumerate(ws.iter_rows(min_row=2),2):
            for c in row: c.border=BDR
            if i%2==0:
                for c in row: c.fill=ALT
        for col in ws.columns:
            w=max((len(str(c.value or "")) for c in col),default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width=min(w+4,42)
    for nm in wb.sheetnames: style(wb[nm])
    if "KPIs" in wb.sheetnames:
        for row in wb["KPIs"].iter_rows(min_row=2):
            for c in row: c.fill=KPI
    if "Anomalies" in wb.sheetnames:
        ws=wb["Anomalies"]; fc,ec=None,None
        for c in ws[1]:
            if c.value=="Flag": fc=c.column
            if c.value=="Explanation": ec=c.column
        if fc:
            for row in ws.iter_rows(min_row=2):
                flag=row[fc-1].value; expl=row[ec-1].value if ec else ""
                fill=EXP if "Expected" in str(expl) else (SPK if flag=="SPIKE" else DIP if flag=="DIP" else ALT)
                for c in row: c.fill=fill
    # Color new columns green in monthly sheets
    added_cols = {"calls_per_agent","unique_rate","efficiency","week_number","flagged","flag_reason",
                  "written_per_call","calls_needed_per_agent","inquiry_rate"}
    for mn in MONTH_ORDER:
        if mn not in wb.sheetnames: continue
        ws=wb[mn]; dc,nc=None,None
        new_col_indices = set()
        for c in ws[1]:
            if str(c.value or "").lower() in added_cols:
                new_col_indices.add(c.column)
                c.fill=PatternFill("solid",fgColor="27500A")
                c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
            if c.value=="day_of_week": dc=c.column
            if c.value=="notes": nc=c.column
        for row in ws.iter_rows(min_row=2):
            day=str(row[dc-1].value or "") if dc else ""
            note=str(row[nc-1].value or "") if nc else ""
            if day in WEEKEND_DAYS: base=WKD
            elif note in HOLIDAY_EVENTS: base=HOL
            elif note in ("payroll_day","payslip_day"): base=PAY
            elif note=="war_absence_notice": base=WAR
            else: base=None
            for c in row:
                if base: c.fill=base
                if c.column in new_col_indices and base is None:
                    c.fill=NEW  # green for new calculated column cells
    wb.save(output_path)

# ── Chart engine ─────────────────────────────────────────────────────
CHART_KEYWORDS = ["chart","graph","plot","show me","visualize","trend","over time","bar","line"]

def wants_chart(q): return any(k in q.lower() for k in CHART_KEYWORDS)

def build_chart(question, summary, sheets, agents):
    q = question.lower()
    if any(w in q for w in ["agent","rep","performer"]):
        if not agents.empty and "calls_per_hour_gross" in agents.columns:
            df = agents[["agent_name","calls_per_hour_gross"]].rename(columns={"agent_name":"Agent","calls_per_hour_gross":"Calls/hr"})
            return "bar", df.set_index("Agent"), "Agent Performance — Calls per Hour"
    if any(w in q for w in ["day","weekday","sunday","monday","busiest day"]):
        day_data = []
        for mn in MONTH_ORDER:
            df = sheets.get(mn)
            if df is None or "inbound_calls" not in df.columns or "day_of_week" not in df.columns: continue
            is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
            is_holiday = df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df), index=df.index)
            work = df[~is_weekend & ~is_holiday]
            for day in ["Sunday","Monday","Tuesday","Wednesday","Thursday"]:
                vals = pd.to_numeric(work[work["day_of_week"]==day]["inbound_calls"],errors="coerce").dropna()
                if not vals.empty: day_data.append({"Day":day,"Avg Calls":round(float(vals.mean()),0)})
        if day_data:
            df = pd.DataFrame(day_data).groupby("Day")["Avg Calls"].mean().reset_index()
            df["Day"] = pd.Categorical(df["Day"],categories=["Sunday","Monday","Tuesday","Wednesday","Thursday"],ordered=True)
            return "bar", df.sort_values("Day").set_index("Day"), "Avg Calls by Day of Week"
    if any(w in q for w in ["inquir","written","vs","both"]):
        if "Total_Inbound_Calls" in summary.columns and "Total_Written_Inquiries" in summary.columns:
            df = summary[["Month","Total_Inbound_Calls","Total_Written_Inquiries"]].copy()
            df["Month"] = df["Month"].str.replace("_"," ")
            return "bar", df.rename(columns={"Total_Inbound_Calls":"Inbound","Total_Written_Inquiries":"Written"}).set_index("Month"), "Calls vs Inquiries"
    if any(w in q for w in ["staff","agent count","headcount"]):
        if "Avg_Agents" in summary.columns:
            df = summary[["Month","Avg_Agents"]].copy(); df["Month"]=df["Month"].str.replace("_"," ")
            return "line", df.rename(columns={"Avg_Agents":"Avg Agents"}).set_index("Month"), "Avg Agents on Shift"
    if "Total_Inbound_Calls" in summary.columns:
        df = summary[["Month","Total_Inbound_Calls"]].copy(); df["Month"]=df["Month"].str.replace("_"," ")
        return "line", df.rename(columns={"Total_Inbound_Calls":"Inbound Calls"}).set_index("Month"), "Monthly Call Volume"
    return None

# ── Ask agent ────────────────────────────────────────────────────────
def ask_agent(context, question, api_key, summary, sheets, agents):
    ok, err = validate_api_key(api_key)
    if not ok: return None, err, None, None

    chart = build_chart(question, summary, sheets, agents) if wants_chart(question) else None

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-opus-4-5", max_tokens=2000,
            system=SYSTEM_PROMPT,
            messages=[{"role":"user","content":f"{context}\n\nUser request: {question}"}]
        )
        raw = resp.content[0].text
        action = parse_action_block(raw)
        clean  = clean_response_text(raw)
        return clean, None, chart, action
    except anthropic.AuthenticationError:
        return None, "Invalid API key. Check the key in the sidebar.", None, None
    except anthropic.RateLimitError:
        return None, "Rate limit reached. Please wait a moment.", None, None
    except anthropic.APIConnectionError:
        return None, "Connection failed. Check your internet.", None, None
    except Exception as e:
        return None, f"Error: {str(e)}", None, None

# ── Session state ────────────────────────────────────────────────────
defaults = {
    "messages":[], "sheets":None, "summary":None, "kpis":None,
    "anomalies":None, "agents":None, "context":None,
    "file_loaded":False, "excel_path":None, "load_error":None,
    "last_chart":None, "action_log":[], "pending_action":None,
}
for k,v in defaults.items():
    if k not in st.session_state: st.session_state[k]=v

# ── Sidebar ──────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuration")
    api_key = st.text_input("Anthropic API Key", type="password", value=os.environ.get("ANTHROPIC_API_KEY",""))
    if api_key:
        ok,err = validate_api_key(api_key)
        color = "#22c55e" if ok else "#ef4444"
        label = "✅ Key format valid" if ok else f"❌ {err}"
        st.markdown(f'<span style="color:{color};font-size:0.78rem;">{label}</span>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 📁 Upload File")
    uploaded = st.file_uploader("Upload Excel file", type=["xlsx","xls"])
    st.markdown("---")
    st.markdown("### 📋 Work Week")
    st.markdown("🟢 **Sun–Thu** = Workdays\n\n🔴 **Fri–Sat** = Weekend")
    st.markdown("---")
    st.markdown("### ✏️ Write Commands")
    st.markdown("""Try:
- *Add a calls per agent column*
- *Add a unique call rate column*
- *Create an anomaly days sheet*
- *Flag all days above 600 calls*
- *Add a weekly total row*
- *Create an executive summary sheet*
- *Add a week number column*""")
    st.markdown("---")
    if st.session_state.action_log:
        st.markdown("### 📋 Action Log")
        for entry in st.session_state.action_log[-5:]:
            st.markdown(f'<div class="action-log-item">{entry}</div>', unsafe_allow_html=True)

# ── Header ───────────────────────────────────────────────────────────
st.markdown("""
<div class="agent-header">
    <div class="status-dot"></div>
    <div>
        <h1>Contact Center AI Agent</h1>
        <p>Hebrew → English · April 2024–September 2025 · Israeli work week (Sun–Thu) · Write actions enabled</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Load file with progress ───────────────────────────────────────────
if uploaded and not st.session_state.file_loaded:
    ok, err = validate_file(uploaded)
    if not ok:
        show_error("File Error", err, "Upload a valid .xlsx or .xls file under 50MB.")
        st.session_state.load_error = err
    else:
        st.session_state.load_error = None
        STEPS = ["Reading Excel file","Translating Hebrew → English","Cleaning data",
                 "Building monthly summary","Computing KPIs","Detecting anomalies",
                 "Analyzing agents","Generating output file"]
        prog = st.empty()
        try:
            with tempfile.NamedTemporaryFile(delete=False,suffix=".xlsx") as tmp:
                tmp.write(uploaded.read()); tmp_path=tmp.name

            show_progress(prog,STEPS,0); sheets=load_and_translate(tmp_path); time.sleep(0.2)
            show_progress(prog,STEPS,1); time.sleep(0.2)
            show_progress(prog,STEPS,2); sheets=clean_sheets(sheets); time.sleep(0.2)
            show_progress(prog,STEPS,3); summary=build_monthly_summary(sheets); time.sleep(0.2)
            show_progress(prog,STEPS,4); kpis=compute_kpis(summary); time.sleep(0.2)
            show_progress(prog,STEPS,5); anomalies=detect_anomalies(sheets); time.sleep(0.2)
            show_progress(prog,STEPS,6); agents=analyze_agents(sheets); time.sleep(0.2)
            show_progress(prog,STEPS,7)
            out = os.path.join(tempfile.gettempdir(),"cc_analyzed.xlsx")
            save_sheets_to_excel(sheets, out, anomalies, kpis)
            time.sleep(0.3)

            prog.markdown("".join([f'<div class="progress-step step-done">✅ {s}</div>' for s in STEPS]), unsafe_allow_html=True)
            time.sleep(0.5); prog.empty()

            context = build_context(sheets,summary,kpis,anomalies,agents,[])
            st.session_state.update({"sheets":sheets,"summary":summary,"kpis":kpis,"anomalies":anomalies,
                                     "agents":agents,"context":context,"excel_path":out,
                                     "file_loaded":True,"messages":[],"action_log":[],"last_chart":None})
            st.success(f"✅ {uploaded.name} loaded — {len(sheets)} sheets, {len(anomalies)} anomalies")
        except ValueError as e:
            prog.empty(); show_error("Data Error",str(e))
        except Exception as e:
            prog.empty(); show_error("Processing Error",str(e),"Try re-uploading. File must not be password-protected.")

elif uploaded is None:
    st.session_state.file_loaded  = False
    st.session_state.load_error   = None

# ── Dashboard ─────────────────────────────────────────────────────────
if st.session_state.file_loaded:
    kpis=st.session_state.kpis; summary=st.session_state.summary

    # KPIs
    st.markdown('<div class="section-title">Key Performance Indicators</div>', unsafe_allow_html=True)
    c1,c2,c3,c4=st.columns(4)
    with c1: st.markdown(f'<div class="kpi-card"><div class="kpi-label">Total Inbound Calls</div><div class="kpi-value">{kpis.get("total_calls",0):,}</div><div class="kpi-sub">Workdays only (Sun–Thu)</div></div>',unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="kpi-card"><div class="kpi-label">Written Inquiries</div><div class="kpi-value">{kpis.get("total_inquiries",0):,}</div><div class="kpi-sub">Apr 2024 – Sep 2025</div></div>',unsafe_allow_html=True)
    with c3:
        peak=kpis.get("peak_month","—").replace("_"," ")
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">Peak Month</div><div class="kpi-value" style="font-size:1.1rem">{peak}</div><div class="kpi-sub">{kpis.get("peak_calls",0):,} calls</div></div>',unsafe_allow_html=True)
    with c4:
        mom=kpis.get("mom_change",0); color="#22c55e" if mom>=0 else "#ef4444"; arrow="↑" if mom>=0 else "↓"
        st.markdown(f'<div class="kpi-card"><div class="kpi-label">MoM Change</div><div class="kpi-value" style="color:{color}">{arrow} {abs(mom)}%</div><div class="kpi-sub">{kpis.get("prev_month","").replace("_"," ")} → {kpis.get("last_month","").replace("_"," ")}</div></div>',unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Charts
    st.markdown('<div class="section-title">Charts</div>', unsafe_allow_html=True)
    t1,t2,t3=st.tabs(["📈 Monthly Volume","📊 Calls vs Inquiries","📅 Day of Week"])
    with t1:
        if "Total_Inbound_Calls" in summary.columns:
            df=summary[["Month","Total_Inbound_Calls"]].copy(); df["Month"]=df["Month"].str.replace("_"," ")
            st.line_chart(df.rename(columns={"Total_Inbound_Calls":"Inbound Calls"}).set_index("Month"),height=220)
    with t2:
        cc=[c for c in ["Total_Inbound_Calls","Total_Written_Inquiries"] if c in summary.columns]
        if cc:
            df=summary[["Month"]+cc].copy(); df["Month"]=df["Month"].str.replace("_"," ")
            df=df.rename(columns={"Total_Inbound_Calls":"Inbound","Total_Written_Inquiries":"Written"})
            st.bar_chart(df.set_index("Month"),height=220)
    with t3:
        day_data=[]
        for mn in MONTH_ORDER:
            df=st.session_state.sheets.get(mn)
            if df is None or "inbound_calls" not in df.columns or "day_of_week" not in df.columns: continue
            is_wknd=df.get("is_weekend",pd.Series([False]*len(df),index=df.index))
            is_hol=df["notes"].isin(HOLIDAY_EVENTS) if "notes" in df.columns else pd.Series([False]*len(df),index=df.index)
            work=df[~is_wknd & ~is_hol]
            for day in ["Sunday","Monday","Tuesday","Wednesday","Thursday"]:
                vals=pd.to_numeric(work[work["day_of_week"]==day]["inbound_calls"],errors="coerce").dropna()
                if not vals.empty: day_data.append({"Day":day,"Avg Calls":round(float(vals.mean()),0)})
        if day_data:
            df=pd.DataFrame(day_data).groupby("Day")["Avg Calls"].mean().reset_index()
            df["Day"]=pd.Categorical(df["Day"],categories=["Sunday","Monday","Tuesday","Wednesday","Thursday"],ordered=True)
            st.bar_chart(df.sort_values("Day").set_index("Day"),height=220)

    st.markdown("<br>", unsafe_allow_html=True)

    # Anomalies + Agents side by side
    col_left,col_right=st.columns([3,2])
    with col_left:
        st.markdown('<div class="section-title">Anomalies Detected</div>', unsafe_allow_html=True)
        anom=st.session_state.anomalies
        if anom is not None and not anom.empty:
            for _,row in anom.head(10).iterrows():
                flag=row.get("Flag",""); expl=str(row.get("Explanation",""))
                badge=(f'<span class="badge-exp">EXPECTED</span>' if "Expected" in expl
                       else f'<span class="badge-spike">SPIKE</span>' if flag=="SPIKE"
                       else f'<span class="badge-dip">DIP</span>')
                st.markdown(f'<div style="background:#0f1117;border:1px solid #1e2130;border-radius:8px;padding:10px 14px;margin-bottom:8px;font-size:0.82rem;">{badge} <span style="color:#9ca3af;font-family:DM Mono,monospace;margin-left:8px;">{row.get("Date","")} · {row.get("Day","")}</span> <span style="color:#6b7280;margin-left:8px;">{row.get("Month","").replace("_"," ")}</span><br><span style="color:#e0e0e0;font-weight:500;">{int(row.get("Inbound_Calls",0)):,} calls</span> <span style="color:#6b7280;font-size:0.75rem;margin-left:6px;">avg {int(row.get("Month_Avg",0)):,} · z={row.get("Z_Score","")}</span><br><span style="color:#9ca3af;font-size:0.75rem;">{expl}</span></div>', unsafe_allow_html=True)
        else:
            st.markdown('<p style="color:#6b7280;font-size:0.85rem;">No anomalies detected.</p>',unsafe_allow_html=True)

    with col_right:
        st.markdown('<div class="section-title">Agent Performance — Calls per Hour</div>', unsafe_allow_html=True)
        agents_df=st.session_state.agents
        if agents_df is not None and not agents_df.empty:
            for _,row in agents_df.iterrows():
                cph=row.get("calls_per_hour_gross",0); rank=int(row.get("Rank",0))
                bar_w=int(min(cph/20*100,100)) if cph else 0
                medal="🥇" if rank==1 else "🥈" if rank==2 else "🥉" if rank==3 else f"#{rank}"
                st.markdown(f'<div style="background:#0f1117;border:1px solid #1e2130;border-radius:8px;padding:10px 14px;margin-bottom:8px;"><div style="display:flex;justify-content:space-between;align-items:center;"><span style="color:#e0e0e0;font-weight:500;">{medal} {row.get("agent_name","")}</span><span style="color:#22c55e;font-family:DM Mono,monospace;font-size:0.8rem;">{cph:.1f} calls/hr</span></div><div style="background:#1e2130;border-radius:3px;height:4px;margin-top:8px;"><div style="background:#22c55e;width:{bar_w}%;height:4px;border-radius:3px;"></div></div><div style="color:#6b7280;font-size:0.72rem;margin-top:4px;">{row.get("avg_handle_time","")} avg handle time</div></div>', unsafe_allow_html=True)
        else:
            st.markdown('<p style="color:#6b7280;font-size:0.85rem;">No agent data found.</p>',unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Download
    if st.session_state.excel_path and os.path.exists(st.session_state.excel_path):
        with open(st.session_state.excel_path,"rb") as f: xl_bytes=f.read()
        st.download_button("⬇️  Download Analyzed Excel File", data=xl_bytes,
            file_name=f"contact_center_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Chat ──────────────────────────────────────────────────────────
    st.markdown('<div class="section-title">Ask the Agent</div>', unsafe_allow_html=True)

    suggestions = [
        "Analyze this dataset",
        "Show me a chart of monthly calls",
        "Add a calls per agent column",
        "Create an anomaly days sheet",
        "Flag days above 600 calls",
        "Executive summary",
    ]
    scols=st.columns(len(suggestions))
    for i,(col,q) in enumerate(zip(scols,suggestions)):
        with col:
            if st.button(q,key=f"s{i}",use_container_width=True):
                st.session_state._suggested=q

    # ── Pending action confirmation ────────────────────────────────────
    if st.session_state.pending_action:
        action = st.session_state.pending_action
        preview = build_action_preview(action)
        st.markdown(f'<div class="preview-box"><div class="preview-title">⚡ Proposed Action — Review before applying</div><div class="preview-detail">{preview}</div></div>', unsafe_allow_html=True)
        ca,cb=st.columns(2)
        with ca:
            if st.button("✅ Apply this change",use_container_width=True,type="primary"):
                sheets,log,err = apply_action(action, st.session_state.sheets, st.session_state.excel_path)
                if err:
                    st.session_state.messages.append({"role":"error","content":err})
                else:
                    st.session_state.sheets = sheets
                    st.session_state.action_log.append(log)
                    # Regenerate file
                    save_sheets_to_excel(sheets, st.session_state.excel_path, st.session_state.anomalies, st.session_state.kpis)
                    # Update context
                    st.session_state.context = build_context(sheets,st.session_state.summary,st.session_state.kpis,
                                                             st.session_state.anomalies,st.session_state.agents,
                                                             st.session_state.action_log)
                    st.session_state.messages.append({"role":"action","content":f"✅ Action applied: {action.get('description','')}\nLog: {log}"})
                st.session_state.pending_action = None
                st.rerun()
        with cb:
            if st.button("❌ Cancel",use_container_width=True):
                st.session_state.messages.append({"role":"assistant","content":"Action cancelled. No changes were made."})
                st.session_state.pending_action = None
                st.rerun()

    # ── Chat history ──────────────────────────────────────────────────
    if st.session_state.messages:
        chat_html = '<div class="chat-outer">'
        for msg in st.session_state.messages:
            r = msg["role"]; c = msg["content"]
            if r=="user":
                chat_html+=f'<div class="chat-msg"><span class="chat-label">You</span><span class="bubble-user">{c}</span></div>'
            elif r=="error":
                chat_html+=f'<div class="chat-msg"><span class="chat-label">Agent</span><span class="bubble-error">⚠️ {c}</span></div>'
            elif r=="action":
                chat_html+=f'<div class="chat-msg"><span class="chat-label">Agent</span><span class="bubble-action">{c}</span></div>'
            else:
                chat_html+=f'<div class="chat-msg"><span class="chat-label">Agent</span><span class="bubble-agent">{c}</span></div>'
        chat_html+='</div>'
        st.markdown(chat_html, unsafe_allow_html=True)

        # Chart from last response
        lc=st.session_state.last_chart
        if lc:
            ct,cdf,ctitle=lc
            st.markdown(f'<div class="section-title">{ctitle}</div>',unsafe_allow_html=True)
            if ct=="line": st.line_chart(cdf,height=280,use_container_width=True)
            else:          st.bar_chart(cdf,height=280,use_container_width=True)

    # ── Input ─────────────────────────────────────────────────────────
    question=st.chat_input("Ask anything or say 'add a column for...' to modify the file")
    if hasattr(st.session_state,"_suggested"):
        question=st.session_state._suggested; del st.session_state._suggested

    if question:
        if not api_key:
            show_error("No API Key","Enter your Anthropic API key in the sidebar.","Get one at console.anthropic.com")
        else:
            st.session_state.messages.append({"role":"user","content":question})
            with st.spinner(""):
                answer,error,chart,action = ask_agent(
                    st.session_state.context, question, api_key,
                    st.session_state.summary, st.session_state.sheets, st.session_state.agents
                )
            if error:
                st.session_state.messages.append({"role":"error","content":error})
                st.session_state.last_chart=None; st.session_state.pending_action=None
            else:
                st.session_state.messages.append({"role":"assistant","content":answer})
                st.session_state.last_chart=chart
                st.session_state.pending_action=action  # show confirm dialog if action present
            st.rerun()

else:
    if st.session_state.get("load_error"):
        show_error("Could not load file",st.session_state.load_error,"Upload a different file.")
    else:
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;color:#6b7280;">
            <div style="font-size:3rem;margin-bottom:1rem;">📊</div>
            <div style="font-size:1.1rem;color:#9ca3af;margin-bottom:0.5rem;">Upload your Excel file to begin</div>
            <div style="font-size:0.82rem;font-family:'DM Mono',monospace;">Hebrew contact center data · Auto-translates · Write actions enabled</div>
            <br>
            <div style="font-size:0.78rem;color:#4b5563;font-family:'DM Mono',monospace;">Supported: .xlsx .xls · Max 50MB</div>
        </div>""", unsafe_allow_html=True)
