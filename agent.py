import pandas as pd
import anthropic
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
#  SCHEMA — Hebrew to English
# ─────────────────────────────────────────────────────────────────

SHEET_TRANSLATION = {
    "אפריל 24":"April_2024","מאי 24":"May_2024","יוני 24":"June_2024",
    "אוגוסט 24":"August_2024","ספט 24":"September_2024","ינואר 25":"January_2025",
    "פברואר 25":"February_2025","מרץ 25":"March_2025","אפריל 25":"April_2025",
    "מאי 25":"May_2025","יוני 25":"June_2025","יולי 25":"July_2025",
    "אוגוסט 25":"August_2025","ספטמבר 25":"September_2025",
    "נציגים":"Agents","חיזוי אוגוסט+ספט":"Forecast_Aug_Sep",
    "חיזוי-תוצאות וסטיות":"Forecast_vs_Actuals",
}

COLUMN_SCHEMA = {
    "תאריך":"date","יום":"day_of_week","פניות כתובות נפתחו":"written_inquiries",
    "שיחות נכנסות":"inbound_calls","הערות":"notes","מספר נציגים":"num_agents",
    "שיחות ללא כפילויות":"unique_calls","פיק":"is_peak_day","עזר":"aux_calls",
    "אחוז":"aux_pct","מגמה פניות":"trend_inquiries","מגמה שיחות":"trend_calls",
    "נציגים צפויים":"expected_agents","מספר שעות":"total_hours",
    "כדי להגיע ליעד":"calls_needed_for_target","חיזוי קודם פניות":"prev_forecast_inquiries",
    "חיזוי קודם שיחות":"prev_forecast_calls","פער חיזוי פניות":"forecast_gap_inquiries",
    "פער חיזוי שיחות":"forecast_gap_calls","נציג":"agent_name",
    "זמן שיחה+החזק":"avg_handle_time","עמידה בשעות":"hours_compliance",
    "שיחות לשעה גולמי":"calls_per_hour_gross","תחזית שיחות ללא כפילות":"forecast_unique_calls",
    "תחזית פניות":"forecast_inquiries","תחזית כפילויות":"forecast_duplicates",
    "סהכ תחזית שיחות":"forecast_total_calls",
    "כמות שיחות שיוכלו לענות הנציגים":"agent_capacity_calls",
    "הפרש":"capacity_gap","חיזוי שיחות סופי":"final_forecast_calls",
    "פניות נתון אמת":"actual_inquiries","שיחות נתון אמת":"actual_calls",
    "סטייה פניות":"deviation_inquiries","סטייה שיחות":"deviation_calls",
}

DAY_TRANSLATION = {
    "יום ראשון":"Sunday","יום שני":"Monday","יום שלישי":"Tuesday",
    "יום רביעי":"Wednesday","יום חמישי":"Thursday",
    "יום שישי":"Friday",    # WEEKEND — no calls
    "שבת":"Saturday",       # WEEKEND — no calls
}

# Israeli work week: Sunday–Thursday
# Friday and Saturday are the weekend — zero calls, excluded from all KPIs
WEEKEND_DAYS = {"Friday", "Saturday"}
WORKDAYS     = {"Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"}

NOTES_TRANSLATION = {
    "חופש":"holiday",
    "נוכחות":"payroll_day",
    "שכר":"payroll_day",
    "שכר,תלוש?":"payroll_day",
    "תלוש":"payslip_day",
    "תקשור היעדרות מלחמה":"war_absence_notice",
    "ט באב":"Tisha_BAv",
    "ערב חג (ללא שיחות)":"holiday_eve_no_calls",
    "חזרה ללימודים":"back_to_school",
    "תחילת לימודים":"start_of_school_year",
    "חזרה מחופש ראש השנה":"return_from_rosh_hashana",
    "52":"week_52_annotation",
    "ממוצע שיחות ללא משה ותמר":"avg_excl_two_agents",
}

HIGH_VOLUME_EVENTS = {
    "payroll_day","payslip_day","war_absence_notice",
    "back_to_school","start_of_school_year","return_from_rosh_hashana","Tisha_BAv",
}
HOLIDAY_EVENTS = {"holiday","holiday_eve_no_calls"}

MONTH_ORDER = [
    "April_2024","May_2024","June_2024","August_2024","September_2024",
    "January_2025","February_2025","March_2025","April_2025",
    "May_2025","June_2025","July_2025","August_2025","September_2025",
]

# ─────────────────────────────────────────────────────────────────
#  1. LOAD & TRANSLATE
# ─────────────────────────────────────────────────────────────────

def load_and_translate(path):
    xl = pd.ExcelFile(path)
    sheets = {}
    for heb in xl.sheet_names:
        eng = SHEET_TRANSLATION.get(heb, heb)
        df = xl.parse(heb)
        df = df.rename(columns={h:e for h,e in COLUMN_SCHEMA.items() if h in df.columns})
        if "day_of_week" in df.columns:
            df["day_of_week"] = df["day_of_week"].map(
                lambda x: DAY_TRANSLATION.get(str(x), x) if pd.notna(x) else x)
        if "notes" in df.columns:
            df["notes"] = df["notes"].map(
                lambda x: NOTES_TRANSLATION.get(str(x).strip(), str(x)) if pd.notna(x) else x)
        if "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
        # Add a clear is_weekend flag
        if "day_of_week" in df.columns:
            df["is_weekend"] = df["day_of_week"].isin(WEEKEND_DAYS)
        sheets[eng] = df
    return sheets

# ─────────────────────────────────────────────────────────────────
#  2. CLEAN
# ─────────────────────────────────────────────────────────────────

def clean_sheets(sheets):
    log = [
        f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "Source: data.xlsx — Israeli contact center (Hebrew translated to English)",
        "Work week: Sunday to Thursday. Friday and Saturday are weekend — excluded from all KPIs.",
        "Business events: payroll_day, payslip_day, war_absence_notice, holiday, back_to_school, etc.",
    ]
    cleaned = {}
    for name, df in sheets.items():
        df = df.copy()
        before = len(df)
        df = df.drop_duplicates()
        if len(df) < before:
            log.append(f"[{name}] Removed {before-len(df)} duplicate rows.")
        for col in ["inbound_calls","written_inquiries","unique_calls","num_agents"]:
            if col in df.columns:
                n = int(df[col].isnull().sum())
                if n:
                    df[col] = df[col].fillna(0)
                    log.append(f"[{name}] Filled {n} missing values in '{col}' with 0.")
        cleaned[name] = df
    return cleaned, log

# ─────────────────────────────────────────────────────────────────
#  3. MONTHLY SUMMARY — workdays only (Sun–Thu, non-holiday)
# ─────────────────────────────────────────────────────────────────

def build_monthly_summary(sheets):
    rows = []
    for mn in MONTH_ORDER:
        df = sheets.get(mn)
        if df is None: continue

        # A true workday = not weekend AND not holiday
        is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
        is_holiday = pd.Series([False]*len(df), index=df.index)
        if "notes" in df.columns:
            is_holiday = df["notes"].isin(HOLIDAY_EVENTS)

        is_workday = ~is_weekend & ~is_holiday
        w = df[is_workday]

        row = {
            "Month":        mn,
            "Workdays":     int(is_workday.sum()),
            "Weekend_Days": int(is_weekend.sum()),
            "Holiday_Days": int(is_holiday.sum()),
        }
        if "notes" in df.columns:
            row["Payroll_Days"]    = int((df["notes"]=="payroll_day").sum())
            row["Payslip_Days"]    = int((df["notes"]=="payslip_day").sum())
            row["War_Notice_Days"] = int((df["notes"]=="war_absence_notice").sum())

        for col, lbl in [
            ("inbound_calls",     "Total_Inbound_Calls"),
            ("written_inquiries", "Total_Written_Inquiries"),
            ("unique_calls",      "Total_Unique_Calls"),
            ("num_agents",        "Avg_Agents_On_Shift"),
        ]:
            if col in w.columns:
                vals = pd.to_numeric(w[col], errors="coerce").dropna()
                row[lbl] = round(float(vals.sum()),0) if lbl.startswith("Total") else round(float(vals.mean()),1)
        rows.append(row)
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────────────────────────
#  4. KPIs
# ─────────────────────────────────────────────────────────────────

def compute_kpis(summary):
    kpis = {}
    if "Total_Inbound_Calls" not in summary.columns: return kpis
    kpis["Grand Total Inbound Calls"]     = int(summary["Total_Inbound_Calls"].sum())
    kpis["Grand Total Written Inquiries"] = int(summary["Total_Written_Inquiries"].sum()) if "Total_Written_Inquiries" in summary.columns else 0
    peak = summary.loc[summary["Total_Inbound_Calls"].idxmax()]
    low  = summary.loc[summary["Total_Inbound_Calls"].idxmin()]
    kpis["Peak Month"]             = peak["Month"]
    kpis["Peak Month Total Calls"] = int(peak["Total_Inbound_Calls"])
    kpis["Lowest Month"]           = low["Month"]
    kpis["Lowest Month Calls"]     = int(low["Total_Inbound_Calls"])
    if len(summary) >= 2:
        last = float(summary["Total_Inbound_Calls"].iloc[-1])
        prev = float(summary["Total_Inbound_Calls"].iloc[-2])
        kpis["MoM Change (last 2 months)"] = f"{round((last-prev)/prev*100,1)}%" if prev else "N/A"
        kpis["Last Month"] = summary["Month"].iloc[-1]
        kpis["Prev Month"] = summary["Month"].iloc[-2]
    if "Avg_Agents_On_Shift" in summary.columns:
        kpis["Avg Agents Per Month"] = round(float(summary["Avg_Agents_On_Shift"].mean()),1)
    if "Payroll_Days" in summary.columns:
        kpis["Total Payroll-Day Events"] = int(summary["Payroll_Days"].sum())
    if "War_Notice_Days" in summary.columns:
        kpis["Total War-Notice Events"]  = int(summary["War_Notice_Days"].sum())
    kpis["Work Week"] = "Sunday to Thursday (Israeli schedule). Friday and Saturday = weekend, excluded from all KPIs."
    return kpis

# ─────────────────────────────────────────────────────────────────
#  5. ANOMALY DETECTION — workdays only
# ─────────────────────────────────────────────────────────────────

def detect_anomalies(sheets):
    anomalies = []
    for mn in MONTH_ORDER:
        df = sheets.get(mn)
        if df is None or "inbound_calls" not in df.columns: continue

        is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
        is_holiday = pd.Series([False]*len(df), index=df.index)
        if "notes" in df.columns:
            is_holiday = df["notes"].isin(HOLIDAY_EVENTS)

        # Only analyse actual workdays
        work = df[~is_weekend & ~is_holiday].copy()
        vals = pd.to_numeric(work["inbound_calls"], errors="coerce").dropna()
        if len(vals) < 5: continue
        mean, std = float(vals.mean()), float(vals.std())
        if std == 0: continue

        for idx in vals.index:
            v = float(vals.loc[idx])
            z = (v - mean) / std
            if abs(z) > 2.0:
                row = work.loc[idx]
                note = str(row.get("notes",""))
                date_val = row.get("date","")
                date_str = date_val.strftime("%Y-%m-%d") if pd.notna(date_val) and hasattr(date_val,"strftime") else str(date_val)[:10]
                explanation = (
                    f"Expected — business event: {note}" if note in HIGH_VOLUME_EVENTS
                    else "Unexplained spike — investigate" if z > 0
                    else "Unexplained dip — investigate"
                )
                anomalies.append({
                    "Month":          mn,
                    "Date":           date_str,
                    "Day":            str(row.get("day_of_week","")),
                    "Inbound_Calls":  v,
                    "Month_Avg":      round(mean, 0),
                    "Z_Score":        round(z, 2),
                    "Flag":           "SPIKE" if z > 0 else "DIP",
                    "Business_Event": note,
                    "Explanation":    explanation,
                })
    return pd.DataFrame(anomalies)

# ─────────────────────────────────────────────────────────────────
#  6. EVENT IMPACT ANALYSIS — workdays only
# ─────────────────────────────────────────────────────────────────

def event_impact_analysis(sheets):
    rows = []
    for mn in MONTH_ORDER:
        df = sheets.get(mn)
        if df is None or "inbound_calls" not in df.columns or "notes" not in df.columns: continue

        is_weekend = df.get("is_weekend", pd.Series([False]*len(df), index=df.index))
        work = df[~is_weekend & ~df["notes"].isin(HOLIDAY_EVENTS)].copy()

        baseline_vals = pd.to_numeric(
            work[~work["notes"].isin(HIGH_VOLUME_EVENTS)]["inbound_calls"], errors="coerce").dropna()
        if len(baseline_vals) == 0: continue
        baseline = float(baseline_vals.mean())

        for event in HIGH_VOLUME_EVENTS:
            event_rows = work[work["notes"] == event]
            if event_rows.empty: continue
            event_vals = pd.to_numeric(event_rows["inbound_calls"], errors="coerce").dropna()
            if event_vals.empty: continue
            avg_event = float(event_vals.mean())
            uplift = round((avg_event - baseline) / baseline * 100, 0) if baseline else 0
            rows.append({
                "Month":                mn,
                "Event":               event,
                "Event_Day_Avg_Calls":  round(avg_event, 0),
                "Normal_Day_Avg_Calls": round(baseline, 0),
                "Uplift_Pct":          f"+{uplift:.0f}%" if uplift >= 0 else f"{uplift:.0f}%",
            })
    return pd.DataFrame(rows)

# ─────────────────────────────────────────────────────────────────
#  7. AGENT PERFORMANCE
# ─────────────────────────────────────────────────────────────────

def analyze_agents(sheets):
    ag = sheets.get("Agents")
    if ag is None or "agent_name" not in ag.columns: return pd.DataFrame()
    cols = [c for c in ["agent_name","avg_handle_time","hours_compliance","calls_per_hour_gross"] if c in ag.columns]
    df = ag[cols].dropna(subset=["agent_name"]).copy()
    if "calls_per_hour_gross" in df.columns:
        df = df.sort_values("calls_per_hour_gross", ascending=False).reset_index(drop=True)
        df.insert(0, "Rank", range(1, len(df)+1))
    return df

# ─────────────────────────────────────────────────────────────────
#  8. CONTEXT FOR CLAUDE
# ─────────────────────────────────────────────────────────────────

def build_context(sheets, summary, kpis, anomalies, agents, impact):
    ctx = [
        "=== CONTACT CENTER OPERATIONS — Israeli contact center, April 2024 to September 2025 ===",
        "Domain: Daily inbound calls, written inquiries, agent staffing, and forecasting.",
        "All Hebrew data translated to English.",
        "",
        "WORK WEEK: Sunday to Thursday (Israeli schedule).",
        "Friday and Saturday are the weekend — no calls, excluded from all KPIs and averages.",
        "",
        "KEY BUSINESS EVENTS IN NOTES COLUMN:",
        "  payroll_day        = day employees receive their paycheck — consistently +30 to 60% call volume",
        "  payslip_day        = payslip published online — inquiry spike +74 to 190%",
        "  war_absence_notice = company published notice about war-related absences — +93% spike",
        "  back_to_school     = start of school year — +210% spike",
        "  holiday            = public holiday or off day — zero calls",
        "",
        f"MONTHLY SUMMARY (workdays only, weekends excluded):\n{summary.to_string(index=False)}",
        f"\nKPIs:\n{json.dumps(kpis, indent=2)}",
    ]
    if not anomalies.empty:
        ctx.append(f"\nANOMALIES ({len(anomalies)} detected, weekends excluded):\n{anomalies.to_string(index=False)}")
    if not agents.empty:
        ctx.append(f"\nAGENT PERFORMANCE:\n{agents.to_string(index=False)}")
    if not impact.empty:
        ctx.append(f"\nEVENT IMPACT vs NORMAL WORKDAYS:\n{impact.to_string(index=False)}")
    return "\n".join(ctx)

def ask_agent(context, question):
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,
        system="""You are an expert contact center operations analyst.
You analyze data from an Israeli contact center translated from Hebrew to English.
Work week is Sunday to Thursday. Friday and Saturday are the weekend — always excluded from averages and KPIs.
Business events like payroll days, payslip days, and war-related notices significantly affect call volumes.
When answering: be specific and quantitative, reference exact months and dates,
explain business events when they cause anomalies, state assumptions clearly.""",
        messages=[{"role":"user","content":f"{context}\n\nQuestion: {question}"}]
    )
    return response.content[0].text

# ─────────────────────────────────────────────────────────────────
#  9. WRITE OUTPUT EXCEL
# ─────────────────────────────────────────────────────────────────

def write_output(sheets, summary, kpis, anomalies, agents, impact, action_log, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for mn in MONTH_ORDER:
            df = sheets.get(mn)
            if df is None: continue
            d = df.copy()
            if "date" in d.columns:
                d["date"] = d["date"].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x,"strftime") else "")
            d.to_excel(writer, sheet_name=mn[:31], index=False)
        summary.to_excel(writer, sheet_name="Monthly_Summary", index=False)
        pd.DataFrame([{"KPI":k,"Value":str(v)} for k,v in kpis.items()]).to_excel(writer, sheet_name="KPIs", index=False)
        if not anomalies.empty:
            anomalies.to_excel(writer, sheet_name="Anomalies", index=False)
        if not agents.empty:
            agents.to_excel(writer, sheet_name="Agent_Performance", index=False)
        if not impact.empty:
            impact.to_excel(writer, sheet_name="Event_Impact_Analysis", index=False)
        fva = sheets.get("Forecast_vs_Actuals")
        if fva is not None:
            fc = [c for c in ["date","day_of_week","forecast_inquiries","final_forecast_calls",
                               "actual_inquiries","actual_calls","deviation_inquiries","deviation_calls"] if c in fva.columns]
            fd = fva[fc].copy()
            if "date" in fd.columns:
                fd["date"] = fd["date"].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x,"strftime") else "")
            fd.to_excel(writer, sheet_name="Forecast_vs_Actuals", index=False)
        pd.DataFrame({"Action_Log":action_log}).to_excel(writer, sheet_name="Action_Log", index=False)

    # Formatting
    wb = load_workbook(output_path)
    HDR     = PatternFill("solid", fgColor="1F4E79")
    HFONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SPIKE   = PatternFill("solid", fgColor="FFAAAA")   # red   — unexplained spike
    DIP     = PatternFill("solid", fgColor="AAC8FF")   # blue  — dip
    EXPL    = PatternFill("solid", fgColor="FFE5A0")   # amber — expected business event
    WEEKEND = PatternFill("solid", fgColor="E0E0E0")   # gray  — weekend rows
    HOLIDAY = PatternFill("solid", fgColor="D0E8FF")   # light blue — holiday rows
    PAYDAY  = PatternFill("solid", fgColor="FFF2CC")   # yellow — payroll/payslip days
    WAR_F   = PatternFill("solid", fgColor="FFD0D0")   # red tint — war notice days
    SUM_C   = PatternFill("solid", fgColor="E2EFDA")   # green — monthly summary
    KPIC    = PatternFill("solid", fgColor="FFF2CC")   # yellow — KPIs
    AGNT    = PatternFill("solid", fgColor="EBF3FB")   # blue  — agents
    EVNT    = PatternFill("solid", fgColor="EDE7F6")   # purple — event impact
    ALT     = PatternFill("solid", fgColor="F7F7F7")
    thin    = Side(style="thin", color="CCCCCC")
    BDR     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style(ws):
        for c in ws[1]: c.fill=HDR; c.font=HFONT; c.alignment=CTR; c.border=BDR
        ws.freeze_panes = "A2"
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            for c in row:
                c.border = BDR
                if i % 2 == 0: c.fill = ALT
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 42)

    for nm in wb.sheetnames: style(wb[nm])

    if "Monthly_Summary" in wb.sheetnames:
        for row in wb["Monthly_Summary"].iter_rows(min_row=2):
            for c in row: c.fill = SUM_C
    if "KPIs" in wb.sheetnames:
        for row in wb["KPIs"].iter_rows(min_row=2):
            for c in row: c.fill = KPIC
    if "Agent_Performance" in wb.sheetnames:
        for row in wb["Agent_Performance"].iter_rows(min_row=2):
            for c in row: c.fill = AGNT
    if "Event_Impact_Analysis" in wb.sheetnames:
        for row in wb["Event_Impact_Analysis"].iter_rows(min_row=2):
            for c in row: c.fill = EVNT

    # Anomalies: amber=expected event, red=unexplained spike, blue=dip
    if "Anomalies" in wb.sheetnames:
        ws = wb["Anomalies"]
        flag_col, expl_col = None, None
        for c in ws[1]:
            if c.value == "Flag":        flag_col = c.column
            if c.value == "Explanation": expl_col = c.column
        if flag_col:
            for row in ws.iter_rows(min_row=2):
                flag = row[flag_col-1].value
                expl = row[expl_col-1].value if expl_col else ""
                fill = EXPL if "Expected" in str(expl) else (SPIKE if flag=="SPIKE" else DIP if flag=="DIP" else ALT)
                for c in row: c.fill = fill

    # Monthly sheets: color each row by day type
    for mn in MONTH_ORDER:
        if mn not in wb.sheetnames: continue
        ws = wb[mn]
        day_col, note_col = None, None
        for c in ws[1]:
            if c.value == "day_of_week": day_col  = c.column
            if c.value == "notes":       note_col = c.column
        for row in ws.iter_rows(min_row=2):
            day  = str(row[day_col-1].value  or "") if day_col  else ""
            note = str(row[note_col-1].value or "") if note_col else ""
            if day in WEEKEND_DAYS:
                fill = WEEKEND                            # gray — weekend, no work
            elif note in HOLIDAY_EVENTS:
                fill = HOLIDAY                            # blue — public holiday
            elif note in ("payroll_day","payslip_day"):
                fill = PAYDAY                             # yellow — payday spike expected
            elif note == "war_absence_notice":
                fill = WAR_F                              # red tint — war notice spike
            else:
                fill = None
            if fill:
                for c in row: c.fill = fill

    wb.save(output_path)
    print(f"\n  Output saved: {output_path}")

# ─────────────────────────────────────────────────────────────────
#  10. CHAT LOOP
# ─────────────────────────────────────────────────────────────────

def chat_loop(context):
    print("\n" + "="*60)
    print("  CONTACT CENTER AI AGENT")
    print("  Work week: Sunday to Thursday | Weekend: Friday & Saturday")
    print("="*60)
    print("\nExample questions:")
    print("  Analyze this dataset")
    print("  Compare May to April")
    print("  Why was February 2025 so high?")
    print("  What was the impact of the war absence notice?")
    print("  Show underperforming agents")
    print("  What is the average load per workday?")
    print("  Create an executive summary")
    print("  quit\n")
    while True:
        question = input("You: ").strip()
        if question.lower() in ("quit","exit","q"):
            print("Agent: Goodbye.")
            break
        if not question: continue
        print("\nAgent: thinking...\n")
        answer = ask_agent(context, question)
        print(f"Agent:\n{answer}\n")
        print("-"*60)

# ─────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    INPUT_PATH  = "data.xlsx"
    OUTPUT_PATH = "contact_center_analyzed.xlsx"

    print("Loading and translating Hebrew workbook...")
    sheets = load_and_translate(INPUT_PATH)

    print("Cleaning data...")
    sheets, action_log = clean_sheets(sheets)

    print("Building monthly summary (Sun-Thu workdays only)...")
    summary = build_monthly_summary(sheets)

    print("Computing KPIs...")
    kpis = compute_kpis(summary)

    print("Detecting anomalies (workdays only)...")
    anomalies = detect_anomalies(sheets)

    print("Analyzing business event impact...")
    impact = event_impact_analysis(sheets)

    print("Analyzing agent performance...")
    agents = analyze_agents(sheets)

    print("Writing output Excel file...")
    write_output(sheets, summary, kpis, anomalies, agents, impact, action_log, OUTPUT_PATH)

    print("\n" + "="*60)
    print("KPI SNAPSHOT")
    print("="*60)
    for k, v in kpis.items():
        print(f"  {k}: {v}")
    print(f"\n  Anomalies: {len(anomalies)}")
    print(f"  Event impact rows: {len(impact)}")

    context = build_context(sheets, summary, kpis, anomalies, agents, impact)
    chat_loop(context)
